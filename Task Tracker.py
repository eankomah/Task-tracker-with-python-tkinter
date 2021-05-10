# Task Management Program
# The program calculates the hours spent on a task.
# It then calculates the amount of money made for each task at 1hour=$5.
# Time entered and information from the program are be stored in an excel file for future referencing.


from tkinter import *
from tkinter import ttk, messagebox
from tkcalendar import DateEntry
from datetime import datetime
import openpyxl
from openpyxl import Workbook
import pathlib

# create excel sheet or check if its already created
wb = pathlib.Path("Task_tracker_Data.xlsx")
if wb.exists():
    pass
else:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Date(Y-M-D)"
    ws["B1"] = "Start Time(H:M:S)"
    ws["C1"] = "End Time(H:M:S"
    ws["D1"] = "Hours Spent(Hrs)"
    ws["E1"] = "Amount Made($)"

    wb.save("Task_tracker_Data.xlsx")


class Tracker:

    def __init__(self, master):
        master.title("Task Tracker App")
        master.resizable(False, False)

        master.configure(background='#4e95b1')

        self.style = ttk.Style()
        self.style.configure('TFrame', background='#4e95b1')
        self.style.configure('TButton', background='#4e95b1')
        self.style.configure('TLabel', background='#4e95b1', font=('Arial', 11))

        # Header Frame
        self.header_frame = ttk.Frame(master)
        self.header_frame.pack()

        ttk.Label(self.header_frame, text="Welcome to Task Tracker App", font=('Arial', 18, 'bold')).grid(row=0, column=0, columnspan=2 , pady=10)

        # Time Frame
        self.time_frame = ttk.Frame(master)
        self.time_frame.pack()

        # Declare label name for the time frame fields
        ttk.Label(self.time_frame, text='Choose date').grid(row=0, column=0, padx=5, pady=5, sticky='sw')

        ttk.Label(self.time_frame, text='Hrs(24hrs)').grid(row=1, column=1, padx=5, sticky='sw')
        ttk.Label(self.time_frame, text='Mins').grid(row=1, column=2, padx=5, sticky='sw')
        ttk.Label(self.time_frame, text='Secs').grid(row=1, column=3, padx=5, sticky='sw')
        ttk.Label(self.time_frame, text='Start Time:').grid(row=2, column=0, padx=5, sticky='sw')
        ttk.Label(self.time_frame, text='End Time:').grid(row=3, column=0, padx=5, sticky='sw')
        ttk.Label(self.time_frame, text='Amount Earned:').grid(row=4, column=0, padx=5, sticky='sw')
        # amount_label = ttk.Label(self.time_frame, text='').grid(row=4, column=1, columnspan=3, padx=5, sticky='sw')

        # Declare input Time frame fields
        self.entry_date = DateEntry(self.time_frame, width=24, background='darkblue', foreground='white', borderwidth=2, font=('Arial', 12))

        self.hr_start = Spinbox(self.time_frame, from_=0, to_=23, width=7, state='readonly', justify=CENTER, font=('Arial', 12))
        self.min_start = Spinbox(self.time_frame, from_=0, to_=59, width=6, state='readonly', justify=CENTER, font=('Arial', 12))
        self.sec_start = Spinbox(self.time_frame, from_=0, to_=59, width=6, state='readonly', justify=CENTER, font=('Arial', 12))

        self.hr_end = Spinbox(self.time_frame, from_=0, to_=23, width=7, state='readonly', justify=CENTER, font=('Arial', 12))
        self.min_end = Spinbox(self.time_frame, from_=0, to_=59, width=6, state='readonly', justify=CENTER, font=('Arial', 12))
        self.sec_end = Spinbox(self.time_frame, from_=0, to_=59, width=6, state='readonly', justify=CENTER, font=('Arial', 12))

        # Display field Time frame widgets
        self.entry_date.grid(row=0, column=1, columnspan=3, padx=5, pady=5)

        self.hr_start.grid(row=2, column=1, padx=5, pady=5)
        self.min_start.grid(row=2, column=2, padx=5, pady=5)
        self.sec_start.grid(row=2, column=3, padx=5, pady=5)

        self.hr_end.grid(row=3, column=1, padx=5, pady=5)
        self.min_end.grid(row=3, column=2, padx=5, pady=5)
        self.sec_end.grid(row=3, column=3, padx=5, pady=5)

        # Declare buttons
        ttk.Button(self.time_frame, text='Save/Export to Excel', command=self.submit).grid(row=5, column=1, padx=5, pady=15, sticky='e')
        ttk.Button(self.time_frame, text='New Task', command=self.reset).grid(row=5, column=2, padx=5, pady=15, sticky='w')

        # Label Frame
        self.label_frame = ttk.Frame(master)
        self.label_frame.pack()

        # declare label for displaying results

        date_label = ttk.Label(self.label_frame, text="")
        date_label.pack(pady=20)

    def submit(self):
        # get individual values from the field inputs
        date = self.entry_date.get_date()
        date_str = date.strftime('%Y-%m-%d')

        hr_start = self.hr_start.get()
        min_start = self.min_start.get()
        sec_start = self.sec_start.get()

        hr_end = self.hr_end.get()
        min_end = self.min_end.get()
        sec_end = self.sec_end.get()

        # assign entries as time value
        s_time = hr_start + ":" + min_start + ":" + sec_start
        e_time = hr_end + ":" + min_end + ":" + sec_end

        start_time = date_str + " " + s_time
        end_time = date_str + " " + e_time

        # convert all entries into datatime stamp
        starting_timestamp = datetime.strptime(start_time, '%Y-%m-%d %H:%M:%S')
        ending_timestamp = datetime.strptime(end_time, '%Y-%m-%d %H:%M:%S')

        # Validate time entries
        if ending_timestamp < starting_timestamp:
            messagebox.showinfo(title="ERROR", message="Invalid Entry for End Time. End Time can't be lower than Start Time")
        else:
            time_spent = ending_timestamp - starting_timestamp

            hrs_spent = time_spent.total_seconds()/3600

            amount_made = hrs_spent*5

        # print(hrs_spent)
        # print(amount_made)

        display_amount_label = ttk.Label(self.time_frame, text="" ,font=('Arial', 15, 'bold'))
        display_amount_label.grid(row=4, column=1, columnspan=3, padx=5, sticky='sw')

        display_amount_label.config(text="You made $" + str(amount_made) + " in " + str(hrs_spent)+"hrs")

        # grab entries and save them in excel
        wb = openpyxl.load_workbook("Task_tracker_Data.xlsx")
        ws = wb.active
        ws.cell(column=1, row=ws.max_row + 1, value=date_str)
        ws.cell(column=2, row=ws.max_row, value=s_time)
        ws.cell(column=3, row=ws.max_row, value=e_time)
        ws.cell(column=4, row=ws.max_row, value=hrs_spent)
        ws.cell(column=5, row=ws.max_row, value=amount_made)

        wb.save("Task_tracker_Data.xlsx")


        messagebox.showinfo(title="Task Saved",
                            message="Your task has been saved and export to excel successfully")

        # Load Workbook and display to user
        # wb = openpyxl.load_workbook('Task_tracker_Data.xlsx')
        #
        # ws= wb.active
        # Task_date = ws['A']
        # Start_Time = ws['B']
        # End_Time = ws['C']
        # Hours_spent = ws['D']
        # Amount_made = ws['E']
        #
        # date_label = ttk.Label(self.label_frame, text="")
        # date_label.pack(pady=20)
        #
        # amount_label = ttk.Label(self.label_frame, text="")
        # amount_label.pack(pady=20)
        #
        # date_list=''
        # for cell in Task_date:
        #     date_list = f'{date_list + str(cell.value)}\n'
        #
        # # date_label.config(text=date_list)
        #
        # amount_list = ''
        # for cell in range(int(amount_made)):
        #     amount_list = f'{amount_list + str(cell.value)}\n'
        #
        # amount_label.config(text=amount_list)





    def reset(self):
        pass
        # self.entry_start_time.delete(0, 'end')
        # self.entry_end_time.delete(0, 'end')


def main():
    root = Tk()
    root.geometry("500x500")
    tracker = Tracker(root)
    root.mainloop()


if __name__ == "__main__":
    main()
